import tkinter as tk
from tkinter import simpledialog, ttk, messagebox, filedialog
import openpyxl
from ttkthemes import ThemedStyle

class PughMatrixApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Pugh Matrix Enhanced")
        style = ThemedStyle(self.root)
        style.set_theme("plastik")
        self.criteria_vars = []
        self.options_vars = []
        self.scale_vars = []
        self.weight_labels = []
        self.comparison_results = {}
        self.comparison_selections = {}
        self.menu = tk.Menu(self.root)
        self.root.config(menu=self.menu)
        self.filemenu = tk.Menu(self.menu, tearoff=0)
        self.filemenu.add_command(label="New", command=self.new_file)
        self.filemenu.add_command(label="Import from Excel", command=self.import_from_excel)
        self.filemenu.add_command(label="Export to Excel", command=self.export_to_excel)
        self.menu.add_cascade(label="File", menu=self.filemenu)
        self.toolsmenu = tk.Menu(self.menu, tearoff=0)
        self.toolsmenu.add_command(label="Add Criteria", command=self.add_multiple_criteria)
        self.toolsmenu.add_command(label="Add Option", command=self.add_multiple_options)
        self.toolsmenu.add_command(label="Pairwise Comparison", command=self.pairwise_comparison, state='disabled')
        self.toolsmenu.add_command(label="Calculate Score", command=self.calculate_score, state='disabled')
        self.menu.add_cascade(label="Tools", menu=self.toolsmenu)
        ttk.Label(self.root, text="Weight").grid(row=1, column=1, padx=5, pady=5)

    def new_file(self):
        self.root.destroy()
        root = tk.Tk()
        app = PughMatrixApp(root)
        root.mainloop()

    def import_from_excel(self):
        filepath = filedialog.askopenfilename(title="Open Excel File", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if filepath:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            for col in range(3, ws.max_column + 1):
                option = ws.cell(row=1, column=col).value
                self.add_option_column(option)
            for row in range(2, ws.max_row + 1):
                criterion = ws.cell(row=row, column=1).value
                if criterion is None or criterion == "":
                    break  # stop if criterion is empty
                weight = ws.cell(row=row, column=2).value
                self.add_criteria_entry(criterion, weight)
                for col, scale_var in enumerate(self.scale_vars[-1], start=0):
                    value = ws.cell(row=row, column=col + 3).value
                    if value is not None:
                        scale_var.set(float(value))
            comparison_ws = wb['Pairwise Comparison']
            for row in range(1, comparison_ws.max_row + 1):
                a = comparison_ws.cell(row=row, column=1).value
                sign = comparison_ws.cell(row=row, column=2).value
                b = comparison_ws.cell(row=row, column=3).value
                if a and sign and b:
                    self.comparison_results[a] = self.comparison_results.get(a, 0) + (sign == '>')
                    self.comparison_results[b] = self.comparison_results.get(b, 0) + (sign == '<')
                    self.comparison_selections[(a, b)] = sign
            self.update_weights()
            self.root.geometry("")


    def export_to_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pugh Matrix"
        for col, option_var in enumerate(self.options_vars, start=3):
            ws.cell(row=1, column=col, value=option_var.get())
        for row, criterion_var in enumerate(self.criteria_vars, start=2):
            ws.cell(row=row, column=1, value=criterion_var.get())
            ws.cell(row=row, column=2, value=self.weight_labels[row - 2].get())
            for col, scale_var in enumerate(self.scale_vars[row - 2], start=3):
                ws.cell(row=row, column=col, value=scale_var.get())
        comparison_ws = wb.create_sheet("Pairwise Comparison")
        for row, (a, b) in enumerate(self.comparison_selections.items(), start=1):
            comparison_ws.cell(row=row, column=1, value=a[0])
            comparison_ws.cell(row=row, column=2, value=b)
            comparison_ws.cell(row=row, column=3, value=a[1])
        for i, score in enumerate(self.scores):
            ws.cell(row=len(self.criteria_vars) + 3, column=i + 3, value=score)
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if filepath:
            wb.save(filepath)
            messagebox.showinfo("Success", f"Data has been saved to '{filepath}'")

    def add_criteria_entry(self, text=None, weight=None):
        row_position = len(self.criteria_vars) + 2
        criteria_var = tk.StringVar(value=text if text else f"Criteria {len(self.criteria_vars) + 1}")
        ttk.Entry(self.root, textvariable=criteria_var).grid(row=row_position, column=0, padx=5, pady=5)
        self.criteria_vars.append(criteria_var)
        weight_label = tk.StringVar(value=weight if weight else "0")
        self.weight_labels.append(weight_label)
        ttk.Label(self.root, textvariable=weight_label).grid(row=row_position, column=1, padx=5, pady=5)
        if len(self.criteria_vars) > 1:
            self.toolsmenu.entryconfig("Pairwise Comparison", state='normal')
        new_scale_vars = []
        for col, _ in enumerate(self.options_vars):
            var = tk.DoubleVar()
            new_scale_vars.append(var)
            self.create_scale(row_position, col, var)
        self.scale_vars.append(new_scale_vars)
        self.root.geometry("")

    def add_multiple_criteria(self):
        number = simpledialog.askinteger("Input", "Enter number of criteria", minvalue=1, maxvalue=100)
        if number:
            for _ in range(number):
                self.add_criteria_entry()

    def add_multiple_options(self):
        number = simpledialog.askinteger("Input", "Enter number of options", minvalue=1, maxvalue=100)
        if number:
            for _ in range(number):
                self.add_option_column()

    def create_scale(self, row, col, var):
        base_col = col * 2 + 2
        scale = tk.Scale(self.root, from_=-5, to=5, orient=tk.HORIZONTAL, variable=var, sliderlength=10)
        scale.grid(row=row, column=base_col, columnspan=2, padx=5, pady=5, sticky='ew')

    def add_option_column(self, text=None):
        col_position = len(self.options_vars) * 2 + 2
        option_var = tk.StringVar(value=text if text else f"Option {len(self.options_vars) + 1}")
        ttk.Entry(self.root, textvariable=option_var).grid(row=1, column=col_position, columnspan=2, padx=5, pady=5)
        self.options_vars.append(option_var)
        self.toolsmenu.entryconfig("Calculate Score", state='normal')
        for row, _ in enumerate(self.criteria_vars):
            var = tk.DoubleVar()
            self.scale_vars[row].append(var)
            self.create_scale(row + 2, len(self.options_vars) - 1, var)
        self.root.geometry("")

    def pairwise_comparison(self):
        self.comparison_vars = []
        self.comparison_window = tk.Toplevel(self.root)
        self.comparison_window.title("Pairwise Comparison")
        comparisons_frame = ttk.Frame(self.comparison_window)
        comparisons_frame.grid(row=0, column=0)
        if len(self.criteria_vars) * (len(self.criteria_vars) - 1) // 2 > 10:
            comparisons_frame.grid_propagate(False)
            comparisons_frame.config(width=600, height=300)
            canvas = tk.Canvas(comparisons_frame)
            scrollbar = ttk.Scrollbar(comparisons_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.config(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            scrollable_frame.bind("<Configure>", lambda e: canvas.config(scrollregion=canvas.bbox("all")))
        else:
            scrollable_frame = comparisons_frame
        row_idx = 0
        for idx, var1 in enumerate(self.criteria_vars[:-1]):
            for var2 in self.criteria_vars[idx+1:]:
                ttk.Label(scrollable_frame, text=var1.get()).grid(row=row_idx, column=0)
                var = tk.StringVar()
                pair = (var1.get(), var2.get())
                if pair in self.comparison_selections:
                    var.set(self.comparison_selections[pair])
                self.comparison_vars.append((var, var1, var2))
                ttk.Radiobutton(scrollable_frame, text="<", variable=var, value="<").grid(row=row_idx, column=1)
                ttk.Radiobutton(scrollable_frame, text="=", variable=var, value="=").grid(row=row_idx, column=2)
                ttk.Radiobutton(scrollable_frame, text=">", variable=var, value=">").grid(row=row_idx, column=3)
                ttk.Label(scrollable_frame, text=var2.get()).grid(row=row_idx, column=4)
                row_idx += 1
        ttk.Button(self.comparison_window, text="Finish", command=self.finish_pairwise_comparison).grid(row=1, column=0, columnspan=6)

    def finish_pairwise_comparison(self):
        self.comparison_results = {var.get(): 0 for var in self.criteria_vars}
        for var, a, b in self.comparison_vars:
            if var.get() == "<":
                self.comparison_results[a.get()] += 1
            elif var.get() == ">":
                self.comparison_results[b.get()] += 1
            self.comparison_selections[(a.get(), b.get())] = var.get()
        self.update_weights()
        self.comparison_window.destroy()

    def update_weights(self):
        total_comparisons = len(self.criteria_vars) * (len(self.criteria_vars) - 1) // 2
        weights = {criterion: count / total_comparisons for criterion, count in self.comparison_results.items()}
        for criterion_var, weight_label in zip(self.criteria_vars, self.weight_labels):
            weight_label.set(f"{weights.get(criterion_var.get(), 0):.2f}")

    def calculate_score(self):
        total_comparisons = len(self.criteria_vars) * (len(self.criteria_vars) - 1) // 2
        weights = {criterion: count / total_comparisons for criterion, count in self.comparison_results.items()}
        self.scores = [sum(var.get() * weights[criterion_var.get()] for var, criterion_var in zip(scale_vars, self.criteria_vars)) 
                  for scale_vars in zip(*self.scale_vars)]
        for i, score in enumerate(self.scores):
            ttk.Label(self.root, text=f"{score:.2f}", font=('Arial', 14)).grid(row=len(self.criteria_vars) + 2, column=i * 2 + 3, padx=5, pady=5)
        self.root.geometry("")

if __name__ == "__main__":
    root = tk.Tk()
    app = PughMatrixApp(root)
    root.mainloop()
